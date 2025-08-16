import openpyxl


def loadWorkbookSheet(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet


def getRowCount(fileName, sheetName):
    sheet = loadWorkbookSheet(fileName, sheetName)
    return sheet.max_row


def getColumnCount(fileName, sheetName):
    sheet = loadWorkbookSheet(fileName, sheetName)
    return sheet.max_column


def readData(fileName, sheetName, rowNum, colNum):
    sheet = loadWorkbookSheet(fileName, sheetName)
    return sheet.cell(rowNum, colNum).value  # always returns string value


def writeData(fileName, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    sheet.cell(rowNum, colNum).value = data
    workbook.save(fileName)


# workbook = openpyxl.load_workbook(excel_file_path)  # to load workbook
# sheet = workbook["LoginTestData"]  # to load a particular sheet in the workbook
#
# rows = sheet.max_row  # to count the number of rows in sheet - 4
# columns = sheet.max_column  # to count the number of columns in sheet - 2
#
# # Reading data from excel sheet
# for r in range(1, rows+1):
#     for c in range(1, columns+1):
#         print(sheet.cell(r,c).value)
#
# # Writing data in excel sheet
# sheet.cell(1,1).value = "Row1 Column1"
